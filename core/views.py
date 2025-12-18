from functools import wraps
from datetime import timedelta
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.conf import settings 

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.db.models import Q, Count, Sum
from django.utils import timezone
from django.http import HttpResponse, JsonResponse
from django.db import IntegrityError, transaction
from django.utils.safestring import mark_safe
from django.urls import reverse

# --- MODELOS ---
from reservas.models import Reserva, RecursoReserva
from inventario.models import Espacio, Recurso
from .models import Area, Carrera

# --- FORMULARIOS ---
from .forms import (
    CustomUserCreationForm,
    AreaForm,
    CarreraForm,
    EditarUsuarioForm,
    RecursoForm
)

# --- API REST (DRF) ---
from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated
from .serializers import UserSerializer, AreaSerializer, CarreraSerializer

User = get_user_model()


# ==============================================================================
# 1) DECORADOR: SOLO ADMIN (o superuser)
# ==============================================================================

def admin_required(view_func):
    """
    Si NO estás logueado => login
    Si estás logueado pero NO eres ADMIN => home con error
    """
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')

        if getattr(request.user, 'rol', '') != 'ADMIN' and not request.user.is_superuser:
            messages.error(request, 'Acceso denegado: No tienes permisos de administrador.')
            return redirect('home')

        return view_func(request, *args, **kwargs)

    return _wrapped_view


# ==============================================================================
# Helper simple (tu proyecto lo usa para Chart.js)
# ==============================================================================

def dates_to_json_list(lista):
    # En tu caso estás pasando listas a template y usando |safe
    return lista


# ==============================================================================
# 2) VISTA HOME (dashboard para solicitante / usuario normal)
# ==============================================================================

@login_required
def home(request):
    if request.user.rol == 'ADMIN':
        return redirect('admin_dashboard')

    hoy = timezone.localdate()

    if request.user.rol == 'SOLICITANTE':
        reservas_hoy = Reserva.objects.filter(solicitante=request.user, fecha=hoy).count()
        pendientes = Reserva.objects.filter(solicitante=request.user, estado='PENDIENTE').count()
        aprobadas = Reserva.objects.filter(solicitante=request.user, estado='APROBADA').count()
        rechazadas = Reserva.objects.filter(solicitante=request.user, estado='RECHAZADA').count()

        ultimas_reservas = Reserva.objects.filter(solicitante=request.user).order_by('-fecha_solicitud')[:5]

        # Gráfico 1: actividad últimos 7 días
        today = timezone.now().date()
        fechas_grafico, datos_grafico = [], []
        for i in range(6, -1, -1):
            fecha = today - timedelta(days=i)
            count = Reserva.objects.filter(solicitante=request.user, fecha=fecha).count()
            fechas_grafico.append(fecha.strftime("%d/%m"))
            datos_grafico.append(count)

        # Gráfico 2: top recursos del usuario
        top_recursos_qs = (
            RecursoReserva.objects.filter(reserva__solicitante=request.user)
            .values('recurso__nombre')
            .annotate(total_pedidos=Sum('cantidad'))
            .order_by('-total_pedidos')[:5]
        )

        recursos_labels = [item['recurso__nombre'] for item in top_recursos_qs]
        recursos_data = [item['total_pedidos'] for item in top_recursos_qs]

        # Gráfico 3: estados del usuario
        estados_labels = dates_to_json_list(["Pendiente", "Aprobada", "Rechazada"])
        estados_data = [pendientes, aprobadas, rechazadas]

    else:
        # Cualquier otro rol: dashboard informativo vacío
        reservas_hoy = pendientes = aprobadas = rechazadas = 0
        ultimas_reservas = []
        fechas_grafico = [timezone.now().strftime("%d/%m")]
        datos_grafico = [0]
        recursos_labels = []
        recursos_data = []
        estados_labels = dates_to_json_list(["Pendiente", "Aprobada", "Rechazada"])
        estados_data = [0, 0, 0]

    context = {
        'kpi_reservas_hoy': reservas_hoy,
        'kpi_pendientes': pendientes,
        'kpi_aprobadas': aprobadas,
        'kpi_rechazadas': rechazadas,
        'ultimas_reservas': ultimas_reservas,
        'kpi_espacios_disponibles': Espacio.objects.filter(activo=True).count(),
        'kpi_recursos': Recurso.objects.count(),

        'chart_labels': dates_to_json_list(fechas_grafico),
        'chart_data': datos_grafico,
        'recursos_labels': dates_to_json_list(recursos_labels),
        'recursos_data': recursos_data,
        'estados_labels': estados_labels,
        'estados_data': estados_data,

        'ultima_sync': timezone.now()
    }
    return render(request, 'core/home.html', context)


# ==============================================================================
# 3) DASHBOARD ADMIN (con KPIs extra por carrera/fecha)
# ==============================================================================

@admin_required
def admin_dashboard(request):
    # KPIs base
    pending_count = Reserva.objects.filter(estado='PENDIENTE').count()
    approved_count = Reserva.objects.filter(estado='APROBADA').count()
    rejected_count = Reserva.objects.filter(estado='RECHAZADA').count()
    total_spaces = Espacio.objects.filter(activo=True).count()

    # Recursos críticos
    critical_resources = Recurso.objects.filter(stock__lte=5).count()

    # Gráfico 1: actividad últimos 7 días
    today = timezone.now().date()
    fechas_grafico, datos_grafico = [], []
    for i in range(6, -1, -1):
        fecha = today - timedelta(days=i)
        count = Reserva.objects.filter(fecha=fecha).count()
        fechas_grafico.append(fecha.strftime("%d/%m"))
        datos_grafico.append(count)

    # Gráfico 2: top recursos más solicitados (por cantidad)
    top_recursos_qs = (
        RecursoReserva.objects
        .values('recurso__nombre')
        .annotate(total_pedidos=Sum('cantidad'))
        .order_by('-total_pedidos')[:5]
    )
    recursos_labels = [item['recurso__nombre'] for item in top_recursos_qs]
    recursos_data = [item['total_pedidos'] for item in top_recursos_qs]

    # Gráfico 3: reservas por Área (considera carrera.area y el area directo legacy)
    areas_qs = (
        Reserva.objects
        .values('solicitante__carrera__area__nombre')
        .exclude(solicitante__carrera__isnull=True)
        .exclude(solicitante__carrera__area__isnull=True)
        .annotate(total=Count('id'))
        .order_by('-total')
    )

    areas_legacy_qs = (
        Reserva.objects
        .values('solicitante__area__nombre')
        .filter(solicitante__carrera__isnull=True)
        .exclude(solicitante__area__isnull=True)
        .annotate(total=Count('id'))
        .order_by('-total')
    )

    areas_map = {}
    for item in areas_qs:
        key = item['solicitante__carrera__area__nombre']
        areas_map[key] = areas_map.get(key, 0) + (item['total'] or 0)

    for item in areas_legacy_qs:
        key = item['solicitante__area__nombre']
        areas_map[key] = areas_map.get(key, 0) + (item['total'] or 0)

    areas_sorted = sorted(areas_map.items(), key=lambda x: x[1], reverse=True)
    areas_labels = [k for k, _ in areas_sorted]
    areas_data = [v for _, v in areas_sorted]

    # KPI: carrera con más reservas
    carrera_mas_reservas = (
        Reserva.objects
        .values('solicitante__carrera__nombre')
        .exclude(solicitante__carrera__isnull=True)
        .annotate(total=Count('id'))
        .order_by('-total')
        .first()
    )
    carrera_mas_reservas_nombre = (carrera_mas_reservas or {}).get('solicitante__carrera__nombre') or "N/A"
    carrera_mas_reservas_total = (carrera_mas_reservas or {}).get('total') or 0

    # KPI: carrera con más consumo de recursos
    carrera_mas_stock = (
        RecursoReserva.objects
        .values('reserva__solicitante__carrera__nombre')
        .exclude(reserva__solicitante__carrera__isnull=True)
        .annotate(total=Sum('cantidad'))
        .order_by('-total')
        .first()
    )
    carrera_mas_stock_nombre = (carrera_mas_stock or {}).get('reserva__solicitante__carrera__nombre') or "N/A"
    carrera_mas_stock_total = int((carrera_mas_stock or {}).get('total') or 0)

    # KPI: fecha con más reservas
    fecha_mas_reservas = (
        Reserva.objects
        .values('fecha')
        .annotate(total=Count('id'))
        .order_by('-total')
        .first()
    )
    fecha_mas_reservas_valor = (fecha_mas_reservas or {}).get('fecha')
    fecha_mas_reservas_total = (fecha_mas_reservas or {}).get('total') or 0

    # KPI: fecha con más consumo de recursos
    fecha_mas_stock = (
        RecursoReserva.objects
        .values('reserva__fecha')
        .annotate(total=Sum('cantidad'))
        .order_by('-total')
        .first()
    )
    fecha_mas_stock_valor = (fecha_mas_stock or {}).get('reserva__fecha')
    fecha_mas_stock_total = int((fecha_mas_stock or {}).get('total') or 0)

    context = {
        'pending_count': pending_count,
        'approved_count': approved_count,
        'rejected_count': rejected_count,
        'total_spaces': total_spaces,
        'critical_resources': critical_resources,

        'chart_labels': dates_to_json_list(fechas_grafico),
        'chart_data': datos_grafico,
        'recursos_labels': dates_to_json_list(recursos_labels),
        'recursos_data': recursos_data,
        'areas_labels': dates_to_json_list(areas_labels),
        'areas_data': areas_data,

        # Nuevos KPIs
        'kpi_carrera_mas_reservas_nombre': carrera_mas_reservas_nombre,
        'kpi_carrera_mas_reservas_total': carrera_mas_reservas_total,
        'kpi_carrera_mas_stock_nombre': carrera_mas_stock_nombre,
        'kpi_carrera_mas_stock_total': carrera_mas_stock_total,
        'kpi_fecha_mas_reservas': fecha_mas_reservas_valor,
        'kpi_fecha_mas_reservas_total': fecha_mas_reservas_total,
        'kpi_fecha_mas_stock': fecha_mas_stock_valor,
        'kpi_fecha_mas_stock_total': fecha_mas_stock_total,
    }
    return render(request, 'administracion/dashboard.html', context)


# ==============================================================================
# 4) EXPORTAR REPORTES
# ==============================================================================

@admin_required
def export_reservas_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    fecha_hoy = timezone.now().strftime("%d-%m-%Y")
    filename = f"Reporte_Reservas_{fecha_hoy}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reservas"

    headers = [
        'ID', 'Solicitante', 'RUT', 'Área', 'Carrera', 'Correo',
        'Espacio', 'Fecha', 'Inicio', 'Fin', 'Estado',
        'Recursos Adicionales', 'Motivo'
    ]
    ws.append(headers)

    # estilo de encabezado
    header_fill = PatternFill(start_color="D71920", end_color="D71920", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    reservas = (
        Reserva.objects
        .select_related(
            "solicitante",
            "solicitante__carrera",
            "solicitante__carrera__area",
            "solicitante__area",
            "espacio",
        )
        .prefetch_related("recursos_asociados__recurso")
        .order_by('-fecha', '-hora_inicio')
    )

    for r in reservas:
        solicitante_nombre = (f"{r.solicitante.first_name} {r.solicitante.last_name}").strip() or r.solicitante.email
        carrera_nombre = r.solicitante.carrera.nombre if getattr(r.solicitante, "carrera", None) else "Sin Carrera"

        recursos_str = ", ".join([f"{rr.cantidad}x {rr.recurso.nombre}" for rr in r.recursos_asociados.all()])

        row = [
            r.id,
            solicitante_nombre,
            r.solicitante.rut or "",
            r.solicitante.nombre_area,
            carrera_nombre,
            r.solicitante.email,
            r.espacio.nombre,
            r.fecha.strftime("%d-%m-%Y") if r.fecha else "",
            r.hora_inicio.strftime("%H:%M") if r.hora_inicio else "",
            r.hora_fin.strftime("%H:%M") if r.hora_fin else "",
            r.get_estado_display(),
            recursos_str if recursos_str else "N/A",
            r.motivo or "",
        ]
        ws.append(row)

    # auto width columnas
    for column_cells in ws.columns:
        length = max(len(str(cell.value) or "") for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    wb.save(response)
    return response


@admin_required
def export_reservas_csv(request):
    qs = (
        Reserva.objects
        .select_related(
            "solicitante",
            "solicitante__carrera",
            "solicitante__carrera__area",
            "solicitante__area",
            "espacio",
        )
        .order_by('-fecha')
    )

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="reservas_export.csv"'
    writer = csv.writer(response)
    writer.writerow(['id', 'solicitante', 'rut', 'area', 'carrera', 'espacio', 'fecha', 'hora', 'estado'])

    for r in qs:
        carrera_nombre = r.solicitante.carrera.nombre if getattr(r.solicitante, "carrera", None) else "Sin Carrera"
        writer.writerow([
            r.id,
            r.solicitante.email,
            r.solicitante.rut or "",
            r.solicitante.nombre_area,
            carrera_nombre,
            r.espacio.nombre,
            r.fecha,
            r.hora_inicio,
            r.estado
        ])
    return response


# ==============================================================================
# 5) GESTIÓN DE RESERVAS (ADMIN)
# ==============================================================================

@admin_required
def gestion_reservas(request):
    estado_filter = request.GET.get('estado', 'TODAS')
    con_archivo = request.GET.get('con_archivo')

    qs = (
        Reserva.objects
        .select_related(
            "solicitante",
            "solicitante__carrera",
            "solicitante__carrera__area",
            "solicitante__area",
            "espacio",
        )
        .prefetch_related("recursos_asociados__recurso")
        .order_by('-fecha', '-hora_inicio')
    )

    if estado_filter != 'TODAS':
        qs = qs.filter(estado=estado_filter)

    if con_archivo == 'si':
        qs = qs.filter(archivo_adjunto__isnull=False).exclude(archivo_adjunto='')

    reservas = list(qs)

    for r in reservas:
        r.adjunto_ok = False

        if r.archivo_adjunto and getattr(r.archivo_adjunto, "name", ""):
            try:
                existe = r.archivo_adjunto.storage.exists(r.archivo_adjunto.name)
                r.adjunto_ok = existe

                # ✅ MEJORA: si no existe, limpiarlo en BD (queda “sin adjunto”)
                if not existe:
                    r.archivo_adjunto = None
                    r.save(update_fields=["archivo_adjunto"])

            except Exception:
                r.adjunto_ok = False

    context = {
        "reservas": reservas,
        "filtro_actual": estado_filter,
        "con_archivo": con_archivo,
    }
    return render(request, "administracion/gestion_reservas.html", context)

@admin_required
def aprobar_reserva(request, reserva_id):
    reserva = get_object_or_404(Reserva, pk=reserva_id)

    if request.method == 'POST':
        action = request.POST.get('action')
        confirmado = request.POST.get('confirmado')

        if action == 'APROBAR':
            # Conflictos con otras ya aprobadas (espacio ocupado)
            conflictos_existentes = Reserva.objects.filter(
                espacio=reserva.espacio, fecha=reserva.fecha, estado='APROBADA'
            ).filter(
                Q(hora_inicio__lt=reserva.hora_fin) & Q(hora_fin__gt=reserva.hora_inicio)
            ).exclude(id=reserva.id)

            if conflictos_existentes.exists():
                messages.error(request, "Error: El espacio ya está ocupado por otra reserva aprobada.")
                return redirect('gestion_reservas')

            # Conflicto con pendientes (competencia)
            competencia = Reserva.objects.filter(
                espacio=reserva.espacio, fecha=reserva.fecha, estado='PENDIENTE'
            ).filter(
                Q(hora_inicio__lt=reserva.hora_fin) & Q(hora_fin__gt=reserva.hora_inicio)
            ).exclude(id=reserva.id)

            # Si hay competencia y aún no confirmaron => aviso con botón
            if competencia.exists() and confirmado != 'si':
                ids_conflictivos = ", ".join([f"#{r.id}" for r in competencia])
                csrf_token = request.POST.get('csrfmiddlewaretoken')

                try:
                    url_aprobar = reverse('aprobar_reserva', args=[reserva.id])
                except Exception:
                    url_aprobar = f"/administracion/reservas/aprobar/{reserva.id}/"

                msg_html = f"""
                    <div class="d-flex align-items-center justify-content-between flex-wrap gap-2">
                        <div>
                            <i class="bi bi-exclamation-triangle-fill fs-4 me-2"></i>
                            <strong>¡Conflicto Detectado!</strong><br>
                            Esta reserva choca con {competencia.count()} solicitudes pendientes (IDs: {ids_conflictivos}).
                            <br><small>Al aprobar esta, las demás serán rechazadas automáticamente.</small>
                        </div>
                        <form method="post" action="{url_aprobar}">
                            <input type="hidden" name="csrfmiddlewaretoken" value="{csrf_token}">
                            <input type="hidden" name="action" value="APROBAR">
                            <input type="hidden" name="confirmado" value="si">
                            <button type="submit" class="btn btn-warning btn-sm text-dark fw-bold border-dark shadow-sm">
                                <i class="bi bi-check-circle-fill me-1"></i> Confirmar y Aprobar
                            </button>
                        </form>
                    </div>
                """
                messages.warning(request, mark_safe(msg_html))
                return redirect('gestion_reservas')

            # Aprobar + validar stock por rango horario
            try:
                with transaction.atomic():
                    for rr in reserva.recursos_asociados.all():
                        ocupados = RecursoReserva.objects.filter(
                            recurso=rr.recurso,
                            reserva__estado='APROBADA',
                            reserva__fecha=reserva.fecha
                        ).filter(
                            Q(reserva__hora_inicio__lt=reserva.hora_fin) & Q(reserva__hora_fin__gt=reserva.hora_inicio)
                        ).aggregate(total=Sum('cantidad'))['total'] or 0

                        if (rr.recurso.stock - ocupados) < rr.cantidad:
                            raise ValidationError(f"Stock insuficiente para {rr.recurso.nombre} en el horario solicitado.")

                    reserva.estado = 'APROBADA'
                    reserva.save()

                    if competencia.exists():
                        motivo_rechazo = f"Sistema: Se aprobó una solicitud prioritaria (ID #{reserva.id})."
                        competencia.update(estado='RECHAZADA', motivo_cancelacion=motivo_rechazo)
                        messages.success(request, f'Reserva #{reserva.id} APROBADA. Conflictos rechazados.')
                    else:
                        messages.success(request, f'Reserva #{reserva.id} APROBADA exitosamente.')

            except ValidationError as e:
                messages.error(request, f'No se pudo aprobar: {e.message}')

        elif action == 'RECHAZAR':
            reserva.estado = 'RECHAZADA'
            reserva.save()
            messages.warning(request, f'Reserva #{reserva.id} RECHAZADA.')

    return redirect('gestion_reservas')


@admin_required
def cancelar_forzosamente(request, reserva_id):
    reserva = get_object_or_404(Reserva, pk=reserva_id)

    if request.method == 'POST':
        motivo = request.POST.get('motivo_cancelacion')
        if reserva.estado != 'FINALIZADA':
            reserva.estado = 'CANCELADA'
            if motivo:
                reserva.motivo_cancelacion = f"Cancelada por Admin: {motivo}"
            reserva.save()
            messages.warning(request, f'Reserva #{reserva_id} cancelada.')

    return redirect('gestion_reservas')


# ==============================================================================
# 6) GESTIÓN DE USUARIOS (ADMIN)
# ==============================================================================

@admin_required
def gestion_usuarios(request):
    q = request.GET.get('q', '').strip()
    users = User.objects.all().order_by('rol', 'email')

    if q:
        users = users.filter(
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q) |
            Q(email__icontains=q) |
            Q(rut__icontains=q)
        )

    return render(request, 'administracion/gestion_usuarios.html', {'users': users})


@admin_required
def gestionar_rol_estado(request, user_id):
    user_target = get_object_or_404(User, pk=user_id)

    if request.method == 'POST':
        action = request.POST.get('action')
        value = request.POST.get('value')

        if action == 'rol':
            user_target.rol = value
            messages.success(request, 'Rol actualizado.')
        elif action == 'toggle_active':
            user_target.is_active = not user_target.is_active
            messages.warning(request, 'Estado actualizado.')

        user_target.save()

    return redirect('gestion_usuarios')


@admin_required
def crear_usuario(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            try:
                user = form.save()
                messages.success(request, f'Usuario {user.email} creado correctamente.')
                return redirect('gestion_usuarios')
            except IntegrityError:
                messages.error(request, 'Error: Usuario ya registrado.')
            except Exception as e:
                messages.error(request, f'Error: {e}')
        else:
            messages.error(request, 'Corrige los errores en el formulario.')
    else:
        form = CustomUserCreationForm()

    return render(request, 'administracion/crear_usuario.html', {'form': form})


@admin_required
def editar_usuario(request, user_id):
    """
    Edita usuario (ADMIN y SOLICITANTE):
    - Guarda área (form)
    - Guarda carrera (select manual name="carrera")
    - Si hay carrera, fuerza área = carrera.area (consistencia)
    """
    usuario_a_editar = get_object_or_404(User, pk=user_id)

    # Para el <select> manual en el template
    carreras = Carrera.objects.all().select_related('area').order_by('area__nombre', 'nombre')

    if request.method == 'POST':
        form = EditarUsuarioForm(request.POST, instance=usuario_a_editar)

        if form.is_valid():
            user = form.save(commit=False)

            # ---- Carrera viene del <select manual> ----
            carrera_id = (request.POST.get("carrera") or "").strip()

            if carrera_id:
                # Si eligieron carrera: asignamos carrera y su área
                carrera_obj = get_object_or_404(Carrera, pk=int(carrera_id))
                user.carrera = carrera_obj
                user.area = carrera_obj.area
            else:
                # Sin carrera: limpiamos carrera y dejamos área según form
                user.carrera = None
                user.area = form.cleaned_data.get('area')

            user.save()
            messages.success(request, 'Usuario actualizado.')
            return redirect('gestion_usuarios')

        messages.error(request, 'Corrige los errores en el formulario.')
    else:
        form = EditarUsuarioForm(instance=usuario_a_editar)

    return render(
        request,
        'administracion/editar_usuario.html',
        {'form': form, 'usuario': usuario_a_editar, 'carreras': carreras}
    )


# ==============================================================================
# 7) ÁREAS Y CARRERAS (ADMIN)
# ==============================================================================

@admin_required
def gestion_areas(request):
    areas = Area.objects.all().order_by('nombre')
    return render(request, 'administracion/gestion_areas.html', {'areas': areas})


@admin_required
def crear_area(request):
    if request.method == 'POST':
        form = AreaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Área creada.')
            return redirect('gestion_areas')
    else:
        form = AreaForm()

    return render(request, 'administracion/crear_area.html', {'form': form})


@admin_required
def editar_area(request, area_id):
    area = get_object_or_404(Area, pk=area_id)

    if request.method == 'POST':
        form = AreaForm(request.POST, instance=area)
        if form.is_valid():
            form.save()
            messages.success(request, 'Área actualizada.')
            return redirect('gestion_areas')
    else:
        form = AreaForm(instance=area)

    return render(request, 'administracion/crear_area.html', {'form': form})


@admin_required
def eliminar_area(request, area_id):
    area = get_object_or_404(Area, pk=area_id)
    area.delete()
    messages.warning(request, 'Área eliminada.')
    return redirect('gestion_areas')


@admin_required
def gestion_carreras(request):
    carreras = Carrera.objects.all().select_related('area').order_by('area__nombre', 'nombre')
    return render(request, 'administracion/gestion_carreras.html', {'carreras': carreras})


@admin_required
def crear_carrera(request):
    if request.method == 'POST':
        form = CarreraForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Carrera creada.')
            return redirect('gestion_carreras')
    else:
        form = CarreraForm()

    return render(request, 'administracion/crear_carrera.html', {'form': form})


@admin_required
def editar_carrera(request, carrera_id):
    carrera = get_object_or_404(Carrera, pk=carrera_id)

    if request.method == 'POST':
        form = CarreraForm(request.POST, instance=carrera)
        if form.is_valid():
            form.save()
            messages.success(request, 'Carrera actualizada.')
            return redirect('gestion_carreras')
    else:
        form = CarreraForm(instance=carrera)

    return render(request, 'administracion/crear_carrera.html', {'form': form})


@admin_required
def eliminar_carrera(request, carrera_id):
    carrera = get_object_or_404(Carrera, pk=carrera_id)
    carrera.delete()
    messages.warning(request, 'Carrera eliminada.')
    return redirect('gestion_carreras')


# ==============================================================================
# 8) INVENTARIO (ADMIN)
# ==============================================================================

@admin_required
def gestion_inventario(request):
    """
    Mantengo tu lógica, pero ojo:
    - Aquí usas recurso.area, si tu modelo Recurso no tiene area, te dará error.
    """
    if request.method == 'POST' and request.POST.get('action') == 'crear':
        try:
            form = RecursoForm(request.POST)
            if form.is_valid():
                recurso = form.save(commit=False)

                # Si tu Recurso tiene campo area y quieres autogestionar:
                if hasattr(recurso, "area") and request.user.area:
                    recurso.area = request.user.area

                recurso.save()
                messages.success(request, "Recurso creado correctamente.")
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")
        except Exception as e:
            messages.error(request, f"Error al crear: {e}")

        return redirect('gestion_inventario')

    espacios = Espacio.objects.all()

    # Filtrado por área solo si el modelo Recurso tiene campo area
    if request.user.is_superuser:
        recursos = Recurso.objects.all().order_by('nombre')
    else:
        if getattr(request.user, "area_id", None) and any(f.name == "area" for f in Recurso._meta.get_fields()):
            recursos = Recurso.objects.filter(area=request.user.area).order_by('nombre')
        else:
            recursos = Recurso.objects.all().order_by('nombre')

    return render(request, 'inventario/recursos.html', {'espacios': espacios, 'recursos': recursos})


@admin_required
def editar_recurso(request, recurso_id):
    recurso = get_object_or_404(Recurso, pk=recurso_id)

    if request.method == 'POST':
        form = RecursoForm(request.POST, instance=recurso)
        if form.is_valid():
            form.save()
            messages.success(request, 'Recurso actualizado.')
            return redirect('gestion_inventario')
    else:
        form = RecursoForm(instance=recurso)

    return render(request, 'inventario/editar_recurso.html', {'form': form, 'recurso': recurso})


@admin_required
def eliminar_recurso(request, recurso_id):
    recurso = get_object_or_404(Recurso, pk=recurso_id)

    try:
        recurso.delete()
        messages.success(request, 'Recurso eliminado.')
    except Exception:
        messages.error(request, 'No se puede eliminar (tiene reservas asociadas).')

    return redirect('gestion_inventario')


@admin_required
def espacio_set_estado(request, espacio_id):
    if request.method == 'POST':
        espacio = get_object_or_404(Espacio, pk=espacio_id)
        espacio.activo = not espacio.activo
        espacio.save()
        messages.success(request, f'Estado de {espacio.nombre} actualizado.')

    return redirect('gestion_inventario')


# ==============================================================================
# 9) API REST (DRF)
# ==============================================================================

class AreaViewSet(viewsets.ModelViewSet):
    queryset = Area.objects.all()
    serializer_class = AreaSerializer
    permission_classes = [IsAuthenticated]


class CarreraViewSet(viewsets.ModelViewSet):
    queryset = Carrera.objects.all()
    serializer_class = CarreraSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = Carrera.objects.all()
        area_id = self.request.query_params.get('area_id')
        if area_id is not None:
            queryset = queryset.filter(area_id=area_id)
        return queryset


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]


# ==============================================================================
# 10) API STOCK EN TIEMPO REAL
# ==============================================================================

@login_required
def api_stock_actual(request):
    """
    Retorna JSON con stock disponible, considerando reservas PENDIENTE/APROBADA en el rango horario.
    Este endpoint NO se rompe si Recurso no tiene campo area.
    """
    fecha_str = request.GET.get('fecha')
    hora_inicio_str = request.GET.get('hora_inicio')
    hora_fin_str = request.GET.get('hora_fin')

    def recurso_tiene_area() -> bool:
        try:
            return any(getattr(f, "name", None) == "area" for f in Recurso._meta.get_fields())
        except Exception:
            return False

    recursos = Recurso.objects.all()

    if (
        request.user.is_authenticated
        and not request.user.is_superuser
        and getattr(request.user, "area_id", None)
        and recurso_tiene_area()
    ):
        recursos = Recurso.objects.filter(area=request.user.area)

    data = []

    for r in recursos.order_by("nombre"):
        ocupados = 0

        if fecha_str and hora_inicio_str and hora_fin_str:
            reservas_conflicto = Reserva.objects.filter(
                fecha=fecha_str,
                estado__in=['PENDIENTE', 'APROBADA']
            ).filter(
                Q(hora_inicio__lt=hora_fin_str) & Q(hora_fin__gt=hora_inicio_str)
            )

            if reservas_conflicto.exists():
                ocupados = (
                    RecursoReserva.objects
                    .filter(reserva__in=reservas_conflicto, recurso=r)
                    .aggregate(total=Sum('cantidad'))['total']
                    or 0
                )

        stock_total = int(r.stock or 0)
        ocupado_total = int(ocupados or 0)
        disponible = stock_total - ocupado_total

        data.append({
            'id': r.id,
            'nombre': r.nombre,
            'total': stock_total,
            'disponible': max(disponible, 0)
        })

    return JsonResponse({'recursos': data})


@login_required
def api_reservas_calendario(request):
    """
    Retorna eventos para FullCalendar (solo reservas aprobadas)
    """
    reservas = Reserva.objects.filter(estado='APROBADA')
    eventos = []

    for r in reservas:
        eventos.append({
            'title': f"Ocupado: {r.espacio.nombre}",
            'start': f"{r.fecha}T{r.hora_inicio}",
            'end': f"{r.fecha}T{r.hora_fin}",
            'color': '#D71920',
            'allDay': False,
        })

    return JsonResponse(eventos, safe=False)
