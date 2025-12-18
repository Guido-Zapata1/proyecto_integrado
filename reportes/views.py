from __future__ import annotations

from io import BytesIO
from datetime import datetime, date, time as time_cls

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Sum, Value, F
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from core.views import admin_required
from reservas.models import Reserva, RecursoReserva


# ==============================================================================
# Helpers (Excel + cálculos)
# ==============================================================================

HEADER_FILL = PatternFill("solid", fgColor="D71920")  # Rojo INACAP
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MESES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

# ✅ Para reportes 1–6: solo APROBADA (así no aparece FINALIZADA si no la usas)
OK_STATES = ["APROBADA"]


def _duration_hours(h_inicio, h_fin) -> float:
    """Calcula duración en horas (float) para métricas."""
    if not h_inicio or not h_fin:
        return 0.0
    base = date(2000, 1, 1)
    dt_ini = datetime.combine(base, h_inicio)
    dt_fin = datetime.combine(base, h_fin)
    seconds = (dt_fin - dt_ini).total_seconds()
    if seconds <= 0:
        return 0.0
    return round(seconds / 3600.0, 2)


def _year_from_request(request) -> int:
    """Lee ?year=YYYY y lo valida."""
    try:
        year = int(request.GET.get("year") or "")
    except ValueError:
        year = timezone.localdate().year
    if year < 2000 or year > 2100:
        year = timezone.localdate().year
    return year


# -----------------------------
# Expresiones BD (Área/Carrera)
# -----------------------------

def _area_expr_reserva():
    # area = carrera.area o area legacy o "Sin Área"
    return Coalesce(
        F("solicitante__carrera__area__nombre"),
        F("solicitante__area__nombre"),
        Value("Sin Área"),
    )


def _area_expr_recurso_reserva():
    return Coalesce(
        F("reserva__solicitante__carrera__area__nombre"),
        F("reserva__solicitante__area__nombre"),
        Value("Sin Área"),
    )


def _carrera_expr_reserva():
    return Coalesce(
        F("solicitante__carrera__nombre"),
        Value("Sin Carrera"),
    )


def _carrera_expr_recurso_reserva():
    return Coalesce(
        F("reserva__solicitante__carrera__nombre"),
        Value("Sin Carrera"),
    )


# -----------------------------
# Excel helpers
# -----------------------------

def _auto_fit(ws, min_w: int = 12, max_w: int = 48) -> None:
    """Ajusta el ancho de columnas según el contenido."""
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[letter].width = max(min_w, min(max_w, max_len + 2))


def _safe_table_name(ws_title: str, header_row: int) -> str:
    """
    Nombre de tabla Excel:
    - Debe empezar con letra
    - Solo letras/números/underscore
    - Debe ser único por workbook
    """
    base = "".join(ch if ch.isalnum() else "_" for ch in ws_title)
    base = base.strip("_") or "Sheet"
    name = f"T_{base}_{header_row}"
    if not name[0].isalpha():
        name = "T_" + name
    return name[:60]


def _write_table(ws, title: str, subtitle: str, columns: list[str], rows: list[list]) -> None:
    """
    Escribe una hoja con:
    - Título + subtítulo
    - Encabezados estilo INACAP
    - Tabla Excel (con flechas de filtro/ordenamiento)
    - Freeze panes
    - Bordes
    """
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)

    ws["A2"] = subtitle
    ws["A2"].font = Font(color="666666")

    ws.append([])

    header_row = ws.max_row + 1
    ws.append(columns)

    # Encabezado
    for c in range(1, len(columns) + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    # Datos
    for r in rows:
        ws.append(r)

    # Estilo celdas
    for row_i in range(header_row + 1, ws.max_row + 1):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=row_i, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Tipos: fechas/horas reales para ordenamiento correcto
            val = cell.value
            if isinstance(val, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm"
            elif isinstance(val, date) and not isinstance(val, datetime):
                cell.number_format = "yyyy-mm-dd"
            elif isinstance(val, time_cls):
                cell.number_format = "hh:mm"

    # Congelar encabezado
    ws.freeze_panes = ws[f"A{header_row + 1}"]

    # Crear Tabla Excel con filtros (flechas) y estilo
    last_col = get_column_letter(len(columns))
    table_ref = f"A{header_row}:{last_col}{ws.max_row}"
    table = Table(displayName=_safe_table_name(ws.title, header_row), ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    _auto_fit(ws)


def _excel_response(wb: Workbook, filename: str) -> HttpResponse:
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _get_total_recursos_reserva(reserva: Reserva) -> int:
    return sum(int(x.cantidad or 0) for x in reserva.recursos_asociados.all())


def _recursos_texto(reserva: Reserva) -> str:
    if not reserva.recursos_asociados.exists():
        return ""
    return ", ".join([f"{x.cantidad}x {x.recurso.nombre}" for x in reserva.recursos_asociados.all()])


def _area_name_from_user(user) -> str:
    if getattr(user, "carrera", None) and getattr(user.carrera, "area", None):
        return user.carrera.area.nombre
    if getattr(user, "area", None):
        return user.area.nombre
    return "Sin Área"


def _carrera_name_from_user(user) -> str:
    if getattr(user, "carrera", None):
        return user.carrera.nombre
    return "Sin Carrera"


# ==============================================================================
# Home Reportes
# ==============================================================================

@login_required
def reportes_home(request):
    # Deja entrar a ADMIN y SOLICITANTE
    if getattr(request.user, "rol", "") not in ["ADMIN", "SOLICITANTE"] and not request.user.is_superuser:
        messages.error(request, "No tienes permisos para ver reportes.")
        return redirect("home")

    return render(request, "reportes/reportes_home.html", {"year": _year_from_request(request)})


# ==============================================================================
# MIS REPORTES (CUALQUIER USUARIO LOGUEADO, INCLUYE ADMIN)
# Filtran por solicitante=request.user, así siempre son "mis datos"
# ==============================================================================

@login_required
def u1_mis_reservas_excel(request):
    year = _year_from_request(request)

    reservas = (
        Reserva.objects.filter(solicitante=request.user, fecha__year=year)
        .select_related("espacio", "solicitante", "solicitante__carrera", "solicitante__carrera__area", "solicitante__area")
        .prefetch_related("recursos_asociados__recurso")
        .order_by("-fecha", "-hora_inicio")
    )

    total = reservas.count()
    total_horas = 0.0
    total_rec = 0
    estado_counts = {k: 0 for k, _ in Reserva.ESTADOS}

    detalle_rows = []
    for r in reservas:
        dur = _duration_hours(r.hora_inicio, r.hora_fin)
        rec_total = _get_total_recursos_reserva(r)

        total_horas += dur
        total_rec += rec_total
        estado_counts[r.estado] = estado_counts.get(r.estado, 0) + 1

        area = _area_name_from_user(r.solicitante)
        carrera = _carrera_name_from_user(r.solicitante)

        detalle_rows.append([
            r.id,
            r.fecha,                 # date real
            r.hora_inicio,           # time real
            r.hora_fin,              # time real
            dur,
            area,
            carrera,
            r.espacio.nombre if r.espacio_id else "",
            r.estado,
            rec_total,
            _recursos_texto(r),
            r.motivo or "",
        ])

    resumen_rows = [
        ["Total reservas", total],
        ["Total horas", round(total_horas, 2)],
        ["Total recursos solicitados", int(total_rec)],
        ["Área", _area_name_from_user(request.user)],
        ["Carrera", _carrera_name_from_user(request.user)],
    ]
    for k, _label in Reserva.ESTADOS:
        resumen_rows.append([f"Reservas {k}", estado_counts.get(k, 0)])

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Resumen"
    _write_table(
        ws1,
        title="U1) Mis Reservas (Resumen)",
        subtitle=f"Año {year} | Usuario: {request.user.email}",
        columns=["Indicador", "Valor"],
        rows=resumen_rows,
    )

    ws2 = wb.create_sheet("Detalle")
    _write_table(
        ws2,
        title="U1) Mis Reservas (Detalle pivot)",
        subtitle="Tabla plana para pivots",
        columns=[
            "ID", "Fecha", "Hora inicio", "Hora fin", "Duración(h)",
            "Área", "Carrera", "Espacio", "Estado",
            "Total recursos", "Detalle recursos", "Motivo"
        ],
        rows=detalle_rows,
    )

    return _excel_response(wb, f"U1_mis_reservas_{year}.xlsx")


@login_required
def u2_mis_recursos_excel(request):
    year = _year_from_request(request)

    qs = (
        RecursoReserva.objects.filter(
            reserva__solicitante=request.user,
            reserva__fecha__year=year,
            reserva__estado__in=OK_STATES,
        )
        .values("recurso__nombre")
        .annotate(cantidad_total=Sum("cantidad"), reservas=Count("reserva", distinct=True))
        .order_by("-cantidad_total", "recurso__nombre")
    )

    total = sum(int(x["cantidad_total"] or 0) for x in qs) or 0
    resumen_rows = []
    for x in qs:
        cant = int(x["cantidad_total"] or 0)
        pct = round((cant / total) * 100, 2) if total else 0
        resumen_rows.append([x["recurso__nombre"], cant, int(x["reservas"] or 0), pct])

    detalle_qs = (
        RecursoReserva.objects.filter(
            reserva__solicitante=request.user,
            reserva__fecha__year=year,
            reserva__estado__in=OK_STATES,
        )
        .select_related("recurso", "reserva", "reserva__espacio", "reserva__solicitante", "reserva__solicitante__carrera", "reserva__solicitante__carrera__area", "reserva__solicitante__area")
        .order_by("-reserva__fecha", "-reserva__hora_inicio")
    )

    detalle_rows = []
    for rr in detalle_qs:
        area = _area_name_from_user(rr.reserva.solicitante)
        carrera = _carrera_name_from_user(rr.reserva.solicitante)

        detalle_rows.append([
            rr.reserva.id,
            rr.reserva.fecha,
            rr.reserva.espacio.nombre if rr.reserva.espacio_id else "",
            area,
            carrera,
            rr.recurso.nombre,
            int(rr.cantidad or 0),
            rr.reserva.estado,
        ])

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Resumen"
    _write_table(
        ws1,
        title="U2) Mis Recursos Solicitados (Resumen)",
        subtitle=f"Año {year} | Estados considerados: {', '.join(OK_STATES)}",
        columns=["Recurso", "Cantidad total", "N° reservas", "% del total"],
        rows=resumen_rows,
    )

    ws2 = wb.create_sheet("Detalle")
    _write_table(
        ws2,
        title="U2) Mis Recursos (Detalle pivot)",
        subtitle="Tabla plana para pivots",
        columns=["ID Reserva", "Fecha", "Espacio", "Área", "Carrera", "Recurso", "Cantidad", "Estado"],
        rows=detalle_rows,
    )

    return _excel_response(wb, f"U2_mis_recursos_{year}.xlsx")


@login_required
def u3_mis_espacios_excel(request):
    year = _year_from_request(request)

    reservas = (
        Reserva.objects.filter(
            solicitante=request.user,
            fecha__year=year,
            estado__in=OK_STATES,
        )
        .select_related("espacio", "solicitante", "solicitante__carrera", "solicitante__carrera__area", "solicitante__area")
        .prefetch_related("recursos_asociados")
        .order_by("-fecha", "-hora_inicio")
    )

    total = reservas.count()
    stats = {}
    detalle_rows = []

    for r in reservas:
        espacio = r.espacio.nombre if r.espacio_id else "Sin espacio"
        dur = _duration_hours(r.hora_inicio, r.hora_fin)
        d = stats.setdefault(espacio, {"reservas": 0, "horas": 0.0})
        d["reservas"] += 1
        d["horas"] += dur

        detalle_rows.append([
            r.id,
            r.fecha,
            r.hora_inicio,
            r.hora_fin,
            dur,
            espacio,
            _get_total_recursos_reserva(r),
        ])

    resumen_rows = []
    for espacio, d in sorted(stats.items(), key=lambda x: (-x[1]["reservas"], x[0])):
        reservas_n = d["reservas"]
        horas = round(d["horas"], 2)
        pct = round((reservas_n / total) * 100, 2) if total else 0
        prom = round(horas / reservas_n, 2) if reservas_n else 0
        resumen_rows.append([espacio, reservas_n, pct, horas, prom])

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Resumen"
    _write_table(
        ws1,
        title="U3) Mis Espacios más usados (Resumen)",
        subtitle=f"Año {year} | Estados considerados: {', '.join(OK_STATES)}",
        columns=["Espacio", "N° reservas", "% del total", "Total horas", "Prom. horas/reserva"],
        rows=resumen_rows,
    )

    ws2 = wb.create_sheet("Detalle")
    _write_table(
        ws2,
        title="U3) Mis Espacios (Detalle pivot)",
        subtitle="Tabla plana para pivots",
        columns=["ID Reserva", "Fecha", "Hora inicio", "Hora fin", "Duración(h)", "Espacio", "Total recursos"],
        rows=detalle_rows,
    )

    return _excel_response(wb, f"U3_mis_espacios_{year}.xlsx")


# ==============================================================================
# ADMIN PACK 8 (2 hojas)
# ==============================================================================

@admin_required
def r1_recursos_global_excel(request):
    year = _year_from_request(request)

    qs = (
        RecursoReserva.objects.filter(
            reserva__fecha__year=year,
            reserva__estado__in=OK_STATES,
        )
        .values("recurso__nombre")
        .annotate(cantidad_total=Sum("cantidad"), reservas=Count("reserva", distinct=True))
        .order_by("-cantidad_total", "recurso__nombre")
    )

    total = sum(int(x["cantidad_total"] or 0) for x in qs) or 0
    resumen_rows = []
    for x in qs:
        cant = int(x["cantidad_total"] or 0)
        pct = round((cant / total) * 100, 2) if total else 0
        resumen_rows.append([x["recurso__nombre"], cant, int(x["reservas"] or 0), pct])

    detalle_qs = (
        RecursoReserva.objects.filter(
            reserva__fecha__year=year,
            reserva__estado__in=OK_STATES,
        )
        .annotate(area=_area_expr_recurso_reserva())
        .annotate(carrera=_carrera_expr_recurso_reserva())
        .select_related("recurso", "reserva", "reserva__espacio", "reserva__solicitante")
        .order_by("-reserva__fecha", "-reserva__hora_inicio")
    )

    detalle_rows = []
    for rr in detalle_qs:
        detalle_rows.append([
            rr.reserva.id,
            rr.reserva.fecha,
            rr.reserva.hora_inicio,
            rr.reserva.hora_fin,
            rr.reserva.espacio.nombre if rr.reserva.espacio_id else "",
            rr.reserva.solicitante.get_full_name() or rr.reserva.solicitante.email,
            rr.area,
            rr.carrera,
            rr.recurso.nombre,
            int(rr.cantidad or 0),
            rr.reserva.estado,
        ])

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Resumen"
    _write_table(
        ws1,
        title="R1) Recursos más solicitados (Global)",
        subtitle=f"Año {year} | Estados: {', '.join(OK_STATES)}",
        columns=["Recurso", "Cantidad total", "N° reservas", "% del total"],
        rows=resumen_rows,
    )

    ws2 = wb.create_sheet("Detalle")
    _write_table(
        ws2,
        title="R1) Recursos (Detalle pivot)",
        subtitle="Tabla plana (ideal para pivots)",
        columns=["ID Reserva", "Fecha", "Hora inicio", "Hora fin", "Espacio", "Solicitante", "Área", "Carrera", "Recurso", "Cantidad", "Estado"],
        rows=detalle_rows,
    )

    return _excel_response(wb, f"R1_recursos_global_{year}.xlsx")


@admin_required
def r2_recursos_por_area_excel(request):
    year = _year_from_request(request)

    qs = (
        RecursoReserva.objects.filter(
            reserva__fecha__year=year,
            reserva__estado__in=OK_STATES,
        )
        .annotate(area=_area_expr_recurso_reserva())
        .annotate(carrera=_carrera_expr_recurso_reserva())
        .values("area", "carrera", "recurso__nombre")
        .annotate(cantidad_total=Sum("cantidad"), reservas=Count("reserva", distinct=True))
        .order_by("area", "carrera", "-cantidad_total", "recurso__nombre")
    )

    tot_area = {}
    for x in qs:
        tot_area[x["area"]] = tot_area.get(x["area"], 0) + int(x["cantidad_total"] or 0)

    resumen_rows = []
    for x in qs:
        area = x["area"]
        cant = int(x["cantidad_total"] or 0)
        pct = round((cant / tot_area[area]) * 100, 2) if tot_area.get(area) else 0
        resumen_rows.append([area, x["carrera"], x["recurso__nombre"], cant, int(x["reservas"] or 0), pct])

    detalle_qs = (
        RecursoReserva.objects.filter(
            reserva__fecha__year=year,
            reserva__estado__in=OK_STATES,
        )
        .annotate(area=_area_expr_recurso_reserva())
        .annotate(carrera=_carrera_expr_recurso_reserva())
        .select_related("recurso", "reserva", "reserva__espacio", "reserva__solicitante")
        .order_by("area", "carrera", "-reserva__fecha", "-reserva__hora_inicio")
    )

    detalle_rows = []
    for rr in detalle_qs:
        detalle_rows.append([
            rr.area,
            rr.carrera,
            rr.reserva.id,
            rr.reserva.fecha,
            rr.reserva.espacio.nombre if rr.reserva.espacio_id else "",
            rr.reserva.solicitante.get_full_name() or rr.reserva.solicitante.email,
            rr.recurso.nombre,
            int(rr.cantidad or 0),
            rr.reserva.estado,
        ])

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Resumen"
    _write_table(
        ws1,
        title="R2) Recursos más solicitados por Área",
        subtitle=f"Año {year} | Estados: {', '.join(OK_STATES)}",
        columns=["Área", "Carrera", "Recurso", "Cantidad total", "N° reservas", "% dentro del Área"],
        rows=resumen_rows,
    )

    ws2 = wb.create_sheet("Detalle")
    _write_table(
        ws2,
        title="R2) Recursos por Área (Detalle pivot)",
        subtitle="Tabla plana (ideal para pivots)",
        columns=["Área", "Carrera", "ID Reserva", "Fecha", "Espacio", "Solicitante", "Recurso", "Cantidad", "Estado"],
        rows=detalle_rows,
    )

    return _excel_response(wb, f"R2_recursos_por_area_{year}.xlsx")


@admin_required
def r3_espacios_global_excel(request):
    year = _year_from_request(request)

    reservas = (
        Reserva.objects.filter(fecha__year=year, estado__in=OK_STATES)
        .select_related("espacio", "solicitante", "solicitante__area", "solicitante__carrera", "solicitante__carrera__area")
        .prefetch_related("recursos_asociados")
        .order_by("-fecha", "-hora_inicio")
    )

    total_reservas = reservas.count()
    space_stats = {}
    detalle_rows = []

    for r in reservas:
        espacio = r.espacio.nombre if r.espacio_id else "Sin espacio"
        dur = _duration_hours(r.hora_inicio, r.hora_fin)
        space = space_stats.setdefault(espacio, {"reservas": 0, "horas": 0.0})
        space["reservas"] += 1
        space["horas"] += dur

        area = _area_name_from_user(r.solicitante)
        carrera = _carrera_name_from_user(r.solicitante)

        detalle_rows.append([
            r.id,
            r.fecha,
            r.hora_inicio,
            r.hora_fin,
            dur,
            espacio,
            r.solicitante.get_full_name() or r.solicitante.email,
            area,
            carrera,
            r.estado,
            _get_total_recursos_reserva(r),
        ])

    resumen_rows = []
    for espacio, d in sorted(space_stats.items(), key=lambda x: (-x[1]["reservas"], x[0])):
        reservas_n = d["reservas"]
        horas = round(d["horas"], 2)
        pct = round((reservas_n / total_reservas) * 100, 2) if total_reservas else 0
        prom = round(horas / reservas_n, 2) if reservas_n else 0
        resumen_rows.append([espacio, reservas_n, pct, horas, prom])

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Resumen"
    _write_table(
        ws1,
        title="R3) Espacios más usados (Global)",
        subtitle=f"Año {year} | Estados: {', '.join(OK_STATES)}",
        columns=["Espacio", "N° reservas", "% del total", "Total horas", "Prom. horas/reserva"],
        rows=resumen_rows,
    )

    ws2 = wb.create_sheet("Detalle")
    _write_table(
        ws2,
        title="R3) Reservas por espacio (Detalle pivot)",
        subtitle="Tabla plana (ideal para pivots)",
        columns=["ID", "Fecha", "Hora inicio", "Hora fin", "Duración(h)", "Espacio", "Solicitante", "Área", "Carrera", "Estado", "Total recursos"],
        rows=detalle_rows,
    )

    return _excel_response(wb, f"R3_espacios_global_{year}.xlsx")


@admin_required
def r4_espacios_por_area_excel(request):
    year = _year_from_request(request)

    reservas = (
        Reserva.objects.filter(fecha__year=year, estado__in=OK_STATES)
        .select_related("espacio", "solicitante", "solicitante__area", "solicitante__carrera", "solicitante__carrera__area")
        .prefetch_related("recursos_asociados")
        .order_by("-fecha", "-hora_inicio")
    )

    area_space = {}
    total_por_area = {}
    detalle_rows = []

    for r in reservas:
        espacio = r.espacio.nombre if r.espacio_id else "Sin espacio"
        dur = _duration_hours(r.hora_inicio, r.hora_fin)
        area = _area_name_from_user(r.solicitante)
        carrera = _carrera_name_from_user(r.solicitante)

        key = (area, espacio)
        d = area_space.setdefault(key, {"reservas": 0, "horas": 0.0})
        d["reservas"] += 1
        d["horas"] += dur
        total_por_area[area] = total_por_area.get(area, 0) + 1

        detalle_rows.append([
            area,
            carrera,
            r.id,
            r.fecha,
            espacio,
            r.hora_inicio,
            r.hora_fin,
            dur,
            r.solicitante.get_full_name() or r.solicitante.email,
            r.estado,
            _get_total_recursos_reserva(r),
        ])

    resumen_rows = []
    for (area, espacio), d in sorted(area_space.items(), key=lambda x: (x[0][0], -x[1]["reservas"], x[0][1])):
        reservas_n = d["reservas"]
        horas = round(d["horas"], 2)
        pct = round((reservas_n / total_por_area.get(area, 1)) * 100, 2) if total_por_area.get(area) else 0
        prom = round(horas / reservas_n, 2) if reservas_n else 0
        resumen_rows.append([area, espacio, reservas_n, pct, horas, prom])

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Resumen"
    _write_table(
        ws1,
        title="R4) Espacios más usados por Área",
        subtitle=f"Año {year} | Estados: {', '.join(OK_STATES)}",
        columns=["Área", "Espacio", "N° reservas", "% dentro del Área", "Total horas", "Prom. horas/reserva"],
        rows=resumen_rows,
    )

    ws2 = wb.create_sheet("Detalle")
    _write_table(
        ws2,
        title="R4) Área/Espacio (Detalle pivot)",
        subtitle="Tabla plana (ideal para pivots)",
        columns=["Área", "Carrera", "ID", "Fecha", "Espacio", "Hora inicio", "Hora fin", "Duración(h)", "Solicitante", "Estado", "Total recursos"],
        rows=detalle_rows,
    )

    return _excel_response(wb, f"R4_espacios_por_area_{year}.xlsx")


@admin_required
def r5_uso_por_area_excel(request):
    year = _year_from_request(request)

    reservas = (
        Reserva.objects.filter(fecha__year=year, estado__in=OK_STATES)
        .select_related("espacio", "solicitante", "solicitante__area", "solicitante__carrera", "solicitante__carrera__area")
        .prefetch_related("recursos_asociados")
    )

    total_reservas = reservas.count()

    recursos_por_area = (
        RecursoReserva.objects.filter(reserva__fecha__year=year, reserva__estado__in=OK_STATES)
        .annotate(area=_area_expr_recurso_reserva())
        .values("area")
        .annotate(total_recursos=Sum("cantidad"))
    )
    recursos_map = {x["area"]: int(x["total_recursos"] or 0) for x in recursos_por_area}

    area_stats = {}
    detalle_rows = []

    for r in reservas:
        area = _area_name_from_user(r.solicitante)
        carrera = _carrera_name_from_user(r.solicitante)
        dur = _duration_hours(r.hora_inicio, r.hora_fin)

        d = area_stats.setdefault(area, {"reservas": 0, "horas": 0.0})
        d["reservas"] += 1
        d["horas"] += dur

        detalle_rows.append([
            r.id,
            r.fecha,
            area,
            carrera,
            r.espacio.nombre if r.espacio_id else "",
            r.hora_inicio,
            r.hora_fin,
            dur,
            r.solicitante.get_full_name() or r.solicitante.email,
            r.estado,
            _get_total_recursos_reserva(r),
        ])

    resumen_rows = []
    for area, d in sorted(area_stats.items(), key=lambda x: (-x[1]["reservas"], x[0])):
        reservas_n = d["reservas"]
        horas = round(d["horas"], 2)
        pct = round((reservas_n / total_reservas) * 100, 2) if total_reservas else 0
        prom = round(horas / reservas_n, 2) if reservas_n else 0
        total_rec = recursos_map.get(area, 0)
        resumen_rows.append([area, reservas_n, pct, horas, prom, total_rec])

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Resumen"
    _write_table(
        ws1,
        title="R5) Uso por Área (ranking + %)",
        subtitle=f"Año {year} | Estados: {', '.join(OK_STATES)}",
        columns=["Área", "N° reservas", "% del total", "Total horas", "Prom. horas/reserva", "Total recursos solicitados"],
        rows=resumen_rows,
    )

    ws2 = wb.create_sheet("Detalle")
    _write_table(
        ws2,
        title="R5) Reservas por Área (Detalle pivot)",
        subtitle="Tabla plana (ideal para pivots)",
        columns=["ID", "Fecha", "Área", "Carrera", "Espacio", "Hora inicio", "Hora fin", "Duración(h)", "Solicitante", "Estado", "Total recursos"],
        rows=detalle_rows,
    )

    return _excel_response(wb, f"R5_uso_por_area_{year}.xlsx")


@admin_required
def r6_tendencia_mensual_por_area_excel(request):
    year = _year_from_request(request)

    reservas = (
        Reserva.objects.filter(fecha__year=year, estado__in=OK_STATES)
        .select_related("solicitante", "solicitante__area", "solicitante__carrera", "solicitante__carrera__area", "espacio")
        .prefetch_related("recursos_asociados")
    )

    reservas_mes = {}
    detalle_rows = []

    for r in reservas:
        area = _area_name_from_user(r.solicitante)
        carrera = _carrera_name_from_user(r.solicitante)
        m = r.fecha.month
        dur = _duration_hours(r.hora_inicio, r.hora_fin)
        total_rec = _get_total_recursos_reserva(r)

        reservas_mes.setdefault(area, [0] * 12)[m - 1] += 1

        detalle_rows.append([
            area,
            carrera,
            MESES[m - 1],
            r.id,
            r.fecha,
            r.espacio.nombre if r.espacio_id else "",
            r.hora_inicio,
            r.hora_fin,
            dur,
            total_rec,
            r.solicitante.get_full_name() or r.solicitante.email,
        ])

    resumen_rows = []
    for area in sorted(reservas_mes.keys()):
        fila = [area] + reservas_mes[area] + [sum(reservas_mes[area])]
        resumen_rows.append(fila)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Resumen"
    _write_table(
        ws1,
        title="R6) Tendencia mensual por Área (N° reservas)",
        subtitle=f"Año {year} | Estados: {', '.join(OK_STATES)}",
        columns=["Área"] + MESES + ["Total"],
        rows=resumen_rows,
    )

    ws2 = wb.create_sheet("Detalle")
    _write_table(
        ws2,
        title="R6) Área/Mes (Detalle pivot)",
        subtitle="Tabla plana (ideal para pivots)",
        columns=["Área", "Carrera", "Mes", "ID", "Fecha", "Espacio", "Hora inicio", "Hora fin", "Duración(h)", "Total recursos", "Solicitante"],
        rows=detalle_rows,
    )

    return _excel_response(wb, f"R6_tendencia_mensual_por_area_{year}.xlsx")


@admin_required
def r7_estados_por_area_excel(request):
    year = _year_from_request(request)

    qs = (
        Reserva.objects.filter(fecha__year=year)
        .annotate(area=_area_expr_reserva())
        .values("area", "estado")
        .annotate(total=Count("id"))
    )

    data = {}
    for x in qs:
        area = x["area"]
        estado = x["estado"]
        data.setdefault(area, {})[estado] = int(x["total"] or 0)

    estados_orden = [k for k, _ in Reserva.ESTADOS]

    resumen_rows = []
    for area in sorted(data.keys()):
        total = sum(data[area].get(e, 0) for e in estados_orden)
        aprob = data[area].get("APROBADA", 0)
        pct_aprob = round((aprob / total) * 100, 2) if total else 0
        fila = [area, total] + [data[area].get(e, 0) for e in estados_orden] + [pct_aprob]
        resumen_rows.append(fila)

    detalle_qs = (
        Reserva.objects.filter(fecha__year=year)
        .select_related("solicitante", "solicitante__area", "solicitante__carrera", "solicitante__carrera__area", "espacio")
        .order_by("-fecha", "-hora_inicio")
    )

    detalle_rows = []
    for r in detalle_qs:
        area = _area_name_from_user(r.solicitante)
        carrera = _carrera_name_from_user(r.solicitante)
        detalle_rows.append([
            r.id,
            r.fecha,
            r.hora_inicio,
            r.hora_fin,
            r.espacio.nombre if r.espacio_id else "",
            r.solicitante.get_full_name() or r.solicitante.email,
            area,
            carrera,
            r.estado,
        ])

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Resumen"
    _write_table(
        ws1,
        title="R7) Estados por Área (% aprobación)",
        subtitle=f"Año {year} | Incluye TODOS los estados",
        columns=["Área", "Total"] + estados_orden + ["% aprobación (sobre total)"],
        rows=resumen_rows,
    )

    ws2 = wb.create_sheet("Detalle")
    _write_table(
        ws2,
        title="R7) Reservas por Área/Estado (Detalle pivot)",
        subtitle="Tabla plana (ideal para pivots)",
        columns=["ID", "Fecha", "Hora inicio", "Hora fin", "Espacio", "Solicitante", "Área", "Carrera", "Estado"],
        rows=detalle_rows,
    )

    return _excel_response(wb, f"R7_estados_por_area_{year}.xlsx")


@admin_required
def r8_auditoria_detallada_excel(request):
    year = _year_from_request(request)

    reservas = (
        Reserva.objects.filter(fecha__year=year)
        .select_related("espacio", "solicitante", "solicitante__area", "solicitante__carrera", "solicitante__carrera__area")
        .prefetch_related("recursos_asociados__recurso")
        .order_by("-fecha", "-hora_inicio")
    )

    estados_orden = [k for k, _ in Reserva.ESTADOS]

    area_counts = {}
    area_horas = {}
    area_tot = {}

    detalle_rows = []
    for r in reservas:
        area = _area_name_from_user(r.solicitante)
        carrera = _carrera_name_from_user(r.solicitante)
        dur = _duration_hours(r.hora_inicio, r.hora_fin)
        rec_total = _get_total_recursos_reserva(r)

        area_tot[area] = area_tot.get(area, 0) + 1
        area_horas[area] = round(area_horas.get(area, 0.0) + dur, 2)
        area_counts.setdefault(area, {})
        area_counts[area][r.estado] = area_counts[area].get(r.estado, 0) + 1

        detalle_rows.append([
            r.id,
            r.fecha,
            r.hora_inicio,
            r.hora_fin,
            dur,
            r.estado,
            r.espacio.nombre if r.espacio_id else "",
            r.solicitante.get_full_name() or r.solicitante.email,
            area,
            carrera,
            rec_total,
            _recursos_texto(r),
            r.motivo or "",
            r.fecha_solicitud if r.fecha_solicitud else None,
        ])

    rr_area = (
        RecursoReserva.objects.filter(reserva__fecha__year=year)
        .annotate(area=_area_expr_recurso_reserva())
        .values("area")
        .annotate(total_recursos=Sum("cantidad"))
    )
    recursos_map = {x["area"]: int(x["total_recursos"] or 0) for x in rr_area}

    resumen_rows = []
    for area in sorted(area_tot.keys()):
        total_area = area_tot.get(area, 0)
        aprob = area_counts.get(area, {}).get("APROBADA", 0)
        pct_aprob = round((aprob / total_area) * 100, 2) if total_area else 0
        fila_estados = [area_counts.get(area, {}).get(e, 0) for e in estados_orden]
        resumen_rows.append([
            area,
            total_area,
            *fila_estados,
            pct_aprob,
            area_horas.get(area, 0.0),
            recursos_map.get(area, 0),
        ])

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Resumen"
    _write_table(
        ws1,
        title="R8) Auditoría - Resumen por Área",
        subtitle=f"Año {year} | Incluye TODOS los estados",
        columns=["Área", "Total"] + estados_orden + ["% aprobación", "Total horas", "Total recursos (cant.)"],
        rows=resumen_rows,
    )

    ws2 = wb.create_sheet("Detalle")
    _write_table(
        ws2,
        title="R8) Auditoría - Detalle pivot",
        subtitle="Tabla plana (ideal para pivots)",
        columns=[
            "ID",
            "Fecha",
            "Hora inicio",
            "Hora fin",
            "Duración (h)",
            "Estado",
            "Espacio",
            "Solicitante",
            "Área",
            "Carrera",
            "Total recursos",
            "Detalle recursos",
            "Motivo/Actividad",
            "Fecha solicitud",
        ],
        rows=detalle_rows,
    )

    return _excel_response(wb, f"R8_auditoria_{year}.xlsx")
